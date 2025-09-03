import os
import pythoncom
import logging
import comtypes.client
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Count, Value,Q
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.http import JsonResponse
from django.db.models.functions import Coalesce
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from datetime import date, timedelta
from docx2pdf import convert as docx2pdf_convert
from django.core.mail import send_mail
import pandas as pd

from .models import Tenant, Document, DocumentType, TenantType, EmailReminderLog
from .forms import DocumentForm, TenantForm, DocumentUploadForm

from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import permission_required
import openpyxl
from django.http import HttpResponse
from io import BytesIO
from django.core.mail import EmailMessage

from django.http import HttpResponse
from django.shortcuts import render

def axes_lockout_response(request, credentials, *args, **kwargs):
    """
    Handles locked-out users for django-axes.
    """
    # Option 1: Render a template
    return render(request, "axes_locked.html", status=403)

    # Option 2: Simple message
    # return HttpResponse("Your account is locked. Try again later.", status=403)



@login_required(login_url='login')
def profile_view(request):
    user = request.user
    password_form = PasswordChangeForm(user=user)

    if request.method == 'POST':
        password_form = PasswordChangeForm(user=user, data=request.POST)
        if password_form.is_valid():
            password_form.save()
            update_session_auth_hash(request, password_form.user)
            messages.success(request, '✅ Password changed successfully!')
            return redirect('profile')
        else:
            messages.error(request, '❌ Please correct the errors below.')

    return render(request, 'files/profile.html', {
        'user': user,
        'password_form': password_form,
    })
# ----------------------------
# AUTHENTICATION
# ----------------------------

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid username or password')
    return render(request, 'files/login.html')

def logout_view(request):
    logout(request)
    return redirect('login')

@login_required(login_url='login')
def update_tenant_email(request, tenant_id):
    if request.method == "POST":
        tenant = get_object_or_404(Tenant, id=tenant_id)
        tenant.email = request.POST.get("email")
        tenant.save()
        messages.success(request, "Tenant email updated successfully.")
    return redirect(request.META.get("HTTP_REFERER", "/"))
# ----------------------------
# PROTECTED DASHBOARD
# ----------------------------

@login_required(login_url='login')
def dashboard(request):
    query = request.GET.get('search', '')
    selected_type = request.GET.get('type_filter', '')

    tenants = Tenant.objects.all()

    if query:
        tenants = tenants.filter(name__icontains=query)

    if selected_type:
        tenants = tenants.filter(tenant_type_fk__code=selected_type)

    paginator = Paginator(tenants, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    tenant_types = TenantType.objects.all()

    return render(request, 'files/dashboard.html', {
        'tenants': page_obj,
        'search_query': query,
        'tenant_types': tenant_types,
        'selected_type': selected_type
    })
# ----------------------------
# FILE CONVERSION UTILS
# ----------------------------
def convert_word_to_pdf(input_path, output_path):
    pythoncom.CoInitialize()  # Start COM
    try:
        input_path = os.path.abspath(input_path)
        output_path = os.path.abspath(output_path)

        docx2pdf_convert(input_path, output_path)  # Convert

        if not os.path.exists(output_path):
            raise Exception("❌ PDF not created.")
    except Exception as e:
        print(f"Error converting {input_path} to PDF: {e}")
        raise
    finally:
        pythoncom.CoUninitialize()  # End COM

def convert_excel_to_pdf(input_path, output_path):
    pythoncom.CoInitialize()
    try:
        excel = comtypes.client.CreateObject('Excel.Application')
        excel.Visible = False
        workbook = excel.Workbooks.Open(input_path)
        workbook.ExportAsFixedFormat(0, output_path)
        workbook.Close(False)
        excel.Quit()
    except Exception as e:
        print(f"Error converting Excel to PDF: {e}")
        raise
    finally:
        pythoncom.CoUninitialize()


# ----------------------------
# TENANTS
# ----------------------------


@login_required(login_url='login')
def tenant_detail(request, tenant_id):
    tenant = get_object_or_404(Tenant, pk=tenant_id)
    documents = Document.objects.filter(tenant=tenant)

    document_display_list = []

    for doc in documents:
        original_file_path = os.path.join(settings.MEDIA_ROOT, doc.file.name)
        original_file_url = doc.file.url
        base, ext = os.path.splitext(original_file_path)
        ext = ext.lower()
        pdf_path = base + '.pdf'
        pdf_url = os.path.splitext(original_file_url)[0] + '.pdf'

        # Check if PDF exists
        if os.path.exists(pdf_path):
            display_path = pdf_url
            file_type = 'pdf'
        else:
            display_path = original_file_url
            file_type = ext.lstrip('.')  # e.g., 'docx'

        # Save original file info if it's a Word doc
        if ext in ['.doc', '.docx']:
            original_doc_info = {
                'original_file_path': original_file_url,
                'original_file_type': ext.lstrip('.')
            }
        else:
            original_doc_info = None

        document_display_list.append({
            'document': doc,
            'display_path': display_path,
            'file_type': file_type,
            'original_doc': original_doc_info
        })

    all_doc_types = DocumentType.objects.all()

    return render(request, 'files/tenant_detail.html', {
        'tenant': tenant,
        'documents': document_display_list,
        'all_doc_types': all_doc_types
    })




# views.py
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, permission_required
from .forms import TenantForm, UnitFormSet

@permission_required('files.add_tenant', raise_exception=True)
@login_required(login_url='login')
def add_tenant(request):
    if request.method == 'POST':
        tenant_form = TenantForm(request.POST)
        formset = UnitFormSet(request.POST)

        if tenant_form.is_valid() and formset.is_valid():
            tenant = tenant_form.save()
            # assign tenant to each unit before saving
            units = formset.save(commit=False)
            for unit in units:
                unit.tenant = tenant
                unit.save()
            return redirect('dashboard')
    else:
        tenant_form = TenantForm()
        formset = UnitFormSet()

    return render(request, 'files/add_tenant.html', {
        'form': tenant_form,
        'formset': formset
    })
    

@permission_required('files.delete_tenant', raise_exception=True)
@login_required(login_url='login')
def delete_tenant(request, tenant_id):
    tenant = get_object_or_404(Tenant, id=tenant_id)
    tenant.delete()
    return redirect('dashboard')


# ----------------------------
# DOCUMENTS
# ----------------------------
@permission_required('files.add_document', raise_exception=True)
@login_required(login_url='login')
def upload_document(request):
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)
        files = request.FILES.getlist('file')
        if form.is_valid():
            tenant = form.cleaned_data['tenant']
            doc_type_fk = form.cleaned_data['doc_type_fk']
            expiry_date = form.cleaned_data['expiry_date']

            for file in files:
                document = Document.objects.create(
                    tenant=tenant,
                    doc_type_fk=doc_type_fk,
                    expiry_date=expiry_date,
                    file=file
                )

                input_path = os.path.join(settings.MEDIA_ROOT, document.file.name)
                input_path = os.path.abspath(input_path)
                base, ext = os.path.splitext(input_path)
                ext = ext.lower()
                output_pdf_path = base + '.pdf'

                try:
                    if ext in ['.doc', '.docx']:
                        # Convert Word to PDF but keep original Word file
                        convert_word_to_pdf(input_path, output_pdf_path)
                except Exception as e:
                    print(f"Error converting {document.file.name} to PDF: {e}")
                    messages.warning(request, f"Failed to convert {document.file.name} to PDF.")

            messages.success(request, "Upload completed. PDFs generated if possible.")
            return redirect('upload_document')
    else:
        form = DocumentForm()

    return render(request, 'files/upload_document.html', {'form': form})


@permission_required('files.add_document', raise_exception=True)
@login_required(login_url='login')
def upload_documents_by_type(request, tenant_id=None):
    # Handle both URL parameter and query string (?tenant_id=...)
    tenant_id = tenant_id or request.GET.get("tenant_id")
    tenant = None

    if tenant_id:
        tenant = get_object_or_404(Tenant, id=tenant_id)

    if request.method == 'POST':
        tenant_id_post = request.POST.get('tenant') or tenant_id
        if not tenant_id_post:
            return JsonResponse({'error': 'Tenant is required'}, status=400)

        try:
            tenant = Tenant.objects.get(id=tenant_id_post)
        except Tenant.DoesNotExist:
            return JsonResponse({'error': 'Invalid tenant'}, status=404)

        total_uploaded = 0
        doc_types = DocumentType.objects.all()

        for doc_type in doc_types:
            files = request.FILES.getlist(f"{doc_type.code}_files[]")
            for f in files:
                document = Document.objects.create(
                    tenant=tenant,
                    doc_type_fk=doc_type,
                    file=f
                )
                total_uploaded += 1

                input_path = os.path.join(settings.MEDIA_ROOT, document.file.name)
                input_path = os.path.abspath(input_path)
                base, ext = os.path.splitext(input_path)
                ext = ext.lower()
                output_pdf_path = base + '.pdf'

                try:
                    if ext in ['.doc', '.docx']:
                        # Convert Word to PDF but keep the original Word file
                        convert_word_to_pdf(input_path, output_pdf_path)
                except Exception as e:
                    print(f"Error converting {document.file.name} to PDF: {e}")

        return JsonResponse({'success': f'{total_uploaded} documents uploaded and converted if possible.'})

    form = DocumentUploadForm()
    tenants = Tenant.objects.all()
    doc_types = [(dt.code, dt.label) for dt in DocumentType.objects.all()]

    return render(request, 'files/upload_documents_by_type.html', {
        'form': form,
        'tenants': tenants,
        'tenant_selected': tenant,
        'doc_types': doc_types
    })




@permission_required('files.delete_document', raise_exception=True)
@login_required(login_url='login')
def delete_document(request, doc_id):
    document = get_object_or_404(Document, id=doc_id)
    tenant_id = document.tenant.id

    if request.method == 'POST':
        # Remove main file from storage
        file_path = os.path.join(settings.MEDIA_ROOT, document.file.name)
        if os.path.exists(file_path):
            os.remove(file_path)

        # Also remove generated PDF if exists
        base, ext = os.path.splitext(file_path)
        pdf_path = base + ".pdf"
        if os.path.exists(pdf_path):
            os.remove(pdf_path)

        # Delete DB entry
        document.delete()
        messages.success(request, f'Document "{document.doc_type_fk.label if document.doc_type_fk else "No Type"}" deleted permanently.')

    return redirect('tenant_detail', tenant_id=tenant_id)


@permission_required('files.change_document', raise_exception=True)
@require_POST
@login_required(login_url='login')
def update_expiry_date(request, doc_id):
    doc = get_object_or_404(Document, id=doc_id)
    new_date = request.POST.get('expiry_date')
    if new_date:
        doc.expiry_date = new_date
        doc.save()
        messages.success(request, f'Expiry date updated for document "{doc.doc_type_fk.label if doc.doc_type_fk else "No Type"}".')
    else:
        messages.error(request, 'Invalid expiry date.')
    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))

@login_required(login_url='login')
@permission_required('files.view_document', raise_exception=True)
def expired_documents(request):
    query = request.GET.get("search", "")
    doc_type_filter = request.GET.get("doc_type", "")
    tenant_type_filter = request.GET.get("tenant_type", "")

    # Base queryset: expired documents
    docs = Document.objects.select_related('tenant', 'tenant__tenant_type_fk', 'doc_type_fk') \
        .filter(expiry_date__lt=timezone.now().date())

    # Apply filters
    if query:
        docs = docs.filter(tenant__name__icontains=query)

    if doc_type_filter:
        docs = docs.filter(doc_type_fk_id=doc_type_filter)

    if tenant_type_filter:
        docs = docs.filter(tenant__tenant_type_fk_id=tenant_type_filter)

    # Pagination
    paginator = Paginator(docs.order_by("-expiry_date"), 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Dropdown options
    doc_types = DocumentType.objects.values_list("id", "label")
    tenant_types = TenantType.objects.values_list("id", "label")

    # Pass a helper for the template to show tenant type correctly
    def get_tenant_type(tenant):
        return tenant.tenant_type_fk.label if tenant.tenant_type_fk else (tenant.tenant_type or "N/A")

    return render(request, "files/expired_documents.html", {
        "page_obj": page_obj,
        "query": query,
        "doc_type_filter": doc_type_filter,
        "tenant_type_filter": tenant_type_filter,
        "doc_types": doc_types,
        "tenant_types": tenant_types,
        "get_tenant_type": get_tenant_type,
    })

# ----------------------------
# ANALYTICS
# ----------------------------

@login_required(login_url='login')
@permission_required('files.can_view_analytics', raise_exception=True)
def analytics(request):
    total_documents = Document.objects.count()
    expired_documents_count = Document.objects.filter(expiry_date__lt=date.today()).count()
    expiring_soon_count = Document.objects.filter(
        expiry_date__range=[date.today(), date.today() + timedelta(days=30)]
    ).count()

    documents_per_type = (
        Document.objects
        .values('doc_type_fk__label')
        .annotate(count=Count('tenant', distinct=True))
    )

    tenants_per_type = Tenant.objects.values(
        tenant_type_label=Coalesce('tenant_type_fk__label', Value('Unknown'))
    ).annotate(count=Count('id'))

    doc_types = [item['doc_type_fk__label'] or 'Unknown' for item in documents_per_type]
    doc_type_values = [item['count'] for item in documents_per_type]

    context = {
        'total_documents': total_documents,
        'expired_documents_count': expired_documents_count,
        'expiring_soon_count': expiring_soon_count,
        'documents_per_type': documents_per_type,
        'tenants_per_type': tenants_per_type,
        'doc_types': doc_types,
        'doc_type_values': doc_type_values
    }
    return render(request, 'files/analytics.html', context)


# ----------------------------
# TENANTS + DOCUMENTS COMBO
# ----------------------------

@login_required(login_url='login')
@permission_required('files.view_document', raise_exception=True)
def tenants_with_documents(request):
    search_query = request.GET.get('search', '')
    selected_doc_type = request.GET.get('doc_type', '')
    tenants = Tenant.objects.all().order_by('name')

    if search_query:
        tenants = tenants.filter(name__icontains=search_query)
    if selected_doc_type:
        tenants = tenants.filter(documents__doc_type_fk__code=selected_doc_type).distinct()

    paginator = Paginator(tenants, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'tenants': page_obj,
        'search_query': search_query,
        'doc_types': DocumentType.objects.all(),
        'selected_doc_type': selected_doc_type,
    }

    return render(request, 'files/tenants_documents.html', context)

@login_required(login_url='login')
@permission_required('files.view_document', raise_exception=True)
def tenant_document_status(request):
    # Get filter params from URL query string
    tenant_type_filter = request.GET.get('tenant_type', '')  # e.g. 'office' or ''
    doc_type_filter = request.GET.get('doc_type', '')        # e.g. 'kra' or ''
    sort_order = request.GET.get('sort', 'asc')              # 'asc' or 'desc'
    missing_docs = request.GET.get('missing_docs')          # Checkbox: '1' if checked

    # Get all tenant types and document types for filter dropdowns
    tenant_types = TenantType.objects.all()
    doc_types = DocumentType.objects.all()
    doc_codes = [dt.code for dt in doc_types]

    # Base tenant queryset
    tenants = Tenant.objects.all()

    # Filter tenants by tenant_type_fk.code if filter applied
    if tenant_type_filter:
        tenants = tenants.filter(tenant_type_fk__code=tenant_type_filter)

    # Sort tenants by name
    if sort_order == 'desc':
        tenants = tenants.order_by('-name')
    else:
        tenants = tenants.order_by('name')

    table_data = []

    for i, tenant in enumerate(tenants, start=1):
        # Map tenant's documents by document code
        tenant_docs = {doc.doc_type_fk.code if doc.doc_type_fk else doc.doc_type: "YES"
                       for doc in tenant.documents.all()}

        # Skip tenant if doc_type_filter is active and:
        # - Checkbox is NOT checked: show only tenants that HAVE the doc
        # - Checkbox IS checked: show only tenants that DO NOT have the doc
        if doc_type_filter:
            if missing_docs == "1" and doc_type_filter in tenant_docs:
                continue  # Skip tenants who have the doc
            elif missing_docs != "1" and doc_type_filter not in tenant_docs:
                continue  # Skip tenants who don't have the doc

        docs_values = [tenant_docs.get(code, "") for code in doc_codes]

        table_data.append({
            'no': i,
            'tenant_name': tenant.name,
            'docs_values': docs_values,
        })

    context = {
        'tenant_types': tenant_types,
        'doc_types': doc_types,
        'doc_codes': doc_codes,
        'table_data': table_data,
        'selected_tenant_type': tenant_type_filter,
        'selected_doc_type': doc_type_filter,
        'sort_order': sort_order,
        'missing_docs': missing_docs,  # Pass checkbox state to template
    }
    return render(request, 'files/tenant_status.html', context)

from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect
from django.core.mail import EmailMessage
from django.contrib import messages
import logging
from .models import Tenant, Document, DocumentType, EmailReminderLog

logger = logging.getLogger(__name__)
@login_required(login_url='login')
@permission_required('files.add_emailreminderlog', raise_exception=True)
def send_email_reminder(request, tenant_id):
    tenant = get_object_or_404(Tenant, id=tenant_id)
    all_doc_types = DocumentType.objects.all()
    tenant_docs = Document.objects.filter(tenant=tenant)

    if not tenant.email:
        messages.error(request, "Tenant does not have an email address set.")
        return redirect('tenant_detail', tenant_id=tenant.id)

    if request.method == "POST":
        subject = request.POST.get("subject", "Document Submission Reminder")
        message_from_textarea = request.POST.get("message", "").strip()
        cc_emails_raw = request.POST.get("cc_emails", "").strip()
        selected_doc_type_ids = request.POST.getlist('doc_types')

        # Map tenant documents by type
        tenant_docs_by_type = {}
        for doc in tenant_docs:
            if doc.doc_type_fk:
                tenant_docs_by_type.setdefault(doc.doc_type_fk.id, []).append(doc)

        if message_from_textarea:
            # User typed a message → send as-is
            final_message = message_from_textarea
            selected_doc_types = all_doc_types.filter(
                id__in=selected_doc_type_ids
            ) if selected_doc_type_ids else all_doc_types
        else:
            # Build separate lists for missing and expired
            missing_docs = []
            expired_docs = []

            if selected_doc_type_ids:
                # Only include selected document types
                selected_doc_types = all_doc_types.filter(id__in=selected_doc_type_ids)
            else:
                # Include all document types with missing or expired docs
                selected_doc_types = []
                for dt in all_doc_types:
                    docs = tenant_docs_by_type.get(dt.id, [])
                    if not docs:
                        missing_docs.append(dt.label)
                        selected_doc_types.append(dt)
                    elif any(doc.is_expired() for doc in docs):
                        expired_docs.append(dt.label)
                        selected_doc_types.append(dt)

            if not missing_docs and not expired_docs and not selected_doc_type_ids:
                messages.info(request, "Tenant has all documents valid. No reminder sent.")
                return redirect('tenant_detail', tenant_id=tenant.id)

            # Compose message listing separately
            doc_lines = []
            if missing_docs:
                doc_lines.append("Missing Documents:")
                for doc in missing_docs:
                    doc_lines.append(f"- {doc}")
                doc_lines.append("")  # Empty line

            if expired_docs:
                doc_lines.append("Expired Documents:")
                for doc in expired_docs:
                    doc_lines.append(f"- {doc}")
                doc_lines.append("")

            final_message = f"Dear {tenant.name},\n\nPlease submit or renew the following documents:\n" \
                            f"{chr(10).join(doc_lines)}\nThank you,\nManagement\n\n"

        # Prepare CC emails
        cc_emails = [email.strip() for email in cc_emails_raw.split(",") if email.strip()]

        try:
            email = EmailMessage(
                subject,
                final_message.replace('\n', '<br>'),
                None,
                [tenant.email],
                cc=cc_emails if cc_emails else None,
            )
            email.content_subtype = "html"
            email.send(fail_silently=False)

            EmailReminderLog.objects.create(
                tenant=tenant,
                email=tenant.email,
                subject=subject,
                status='Success',
                message=f"Docs reminded: {', '.join([dt.label for dt in selected_doc_types])} | "
                        f"CC: {', '.join(cc_emails) if cc_emails else 'None'}"
            )
            messages.success(request, f"Reminder email sent to {tenant.name}.")

        except Exception as e:
            logger.error(f"Error sending email to {tenant.email}: {e}", exc_info=True)
            EmailReminderLog.objects.create(
                tenant=tenant,
                email=tenant.email,
                subject=subject,
                status='Failed',
                message=str(e)
            )
            messages.error(request, f"Failed to send email: {e}")

        return redirect('tenant_detail', tenant_id=tenant.id)

    messages.error(request, "Invalid request.")
    return redirect('tenant_detail', tenant_id=tenant.id)




from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from io import BytesIO
import openpyxl
import os
from .models import Document

@login_required(login_url='login')
@permission_required('files.view_document', raise_exception=True)
def export_expired_documents_excel(request):
    # Get filters from query params
    query = request.GET.get('search', '')
    doc_type_filter = request.GET.get('doc_type', '')
    tenant_type_filter = request.GET.get('tenant_type', '')

    # Base queryset: expired documents
    expired_docs = Document.objects.filter(expiry_date__lt=timezone.now()).select_related('tenant', 'doc_type_fk', 'tenant__tenant_type_fk')

    # Apply search filter
    if query:
        expired_docs = expired_docs.filter(tenant__name__icontains=query)

    # Apply document type filter
    if doc_type_filter:
        expired_docs = expired_docs.filter(doc_type_fk_id=doc_type_filter)

    # Apply tenant type filter
    if tenant_type_filter:
        expired_docs = expired_docs.filter(tenant__tenant_type_fk_id=tenant_type_filter)

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expired Documents"

    # Header row
    ws.append(["Tenant Name", "Tenant Type", "Document Type", "Expiry Date", "Email", "File Name"])

    # Data rows
    for doc in expired_docs:
        ws.append([
            doc.tenant.name,
            doc.tenant.tenant_type_fk.label if doc.tenant.tenant_type_fk else 'N/A',
            doc.doc_type_fk.label if doc.doc_type_fk else 'N/A',
            doc.expiry_date.strftime('%Y-%m-%d') if doc.expiry_date else 'N/A',
            doc.tenant.email or '',
            os.path.basename(doc.file.name) if doc.file else ''
        ])

    # Save the workbook to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Return response
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=expired_documents.xlsx'
    return response



@login_required(login_url='login')
@permission_required('files.add_emailreminderlog', raise_exception=True)
def send_reminders_to_all_tenants(request):
    if request.method == "POST":
        subject = request.POST.get("subject", "Document Submission Reminder")
        base_message = request.POST.get("message", "")
        cc_emails_raw = request.POST.get("cc_emails", "").strip()
        selected_doc_type_ids = request.POST.getlist('doc_types')
        all_doc_types = DocumentType.objects.all()

        # If no doc types selected, use ALL document types by default
        if selected_doc_type_ids:
            selected_doc_types = all_doc_types.filter(id__in=selected_doc_type_ids)
        else:
            selected_doc_types = all_doc_types

        cc_emails = [email.strip() for email in cc_emails_raw.split(",") if email.strip()]

        tenants = Tenant.objects.filter(email__isnull=False).exclude(email='')

        for tenant in tenants:
            tenant_docs = Document.objects.filter(tenant=tenant)
            
            # Group tenant docs by doc_type id
            tenant_docs_by_type = {}
            for doc in tenant_docs:
                if doc.doc_type_fk:
                    tenant_docs_by_type.setdefault(doc.doc_type_fk.id, []).append(doc)

            missing_docs_labels = []
            expired_docs_labels = []

            for dt in selected_doc_types:
                docs_of_type = tenant_docs_by_type.get(dt.id, [])
                if not docs_of_type:
                    missing_docs_labels.append(dt.label)
                else:
                    all_expired = all(doc.is_expired() for doc in docs_of_type)
                    if all_expired:
                        expired_docs_labels.append(dt.label)

            if not missing_docs_labels and not expired_docs_labels:
                # Skip sending email if tenant has all valid docs for selected types
                continue

            extra_info = (
                f"Missing documents: {', '.join(missing_docs_labels) if missing_docs_labels else 'None'}\n"
                f"Expired documents: {', '.join(expired_docs_labels) if expired_docs_labels else 'None'}"
            )

            insert_after = "Please submit or renew the following documents:"
            if insert_after in base_message:
                final_message = base_message.replace(insert_after, f"{insert_after}\n{extra_info}")
            else:
                final_message = base_message + "\n\n" + extra_info

            # Insert personalized greeting
            final_message = f"Dear {tenant.name},\n\n" + final_message

            try:
                email = EmailMessage(
                    subject,
                    final_message.replace('\n', '<br>'),
                    None,
                    [tenant.email],
                    cc=cc_emails if cc_emails else None,
                )
                email.content_subtype = "html"
                email.send(fail_silently=False)

                EmailReminderLog.objects.create(
                    tenant=tenant,
                    email=tenant.email,
                    subject=subject,
                    status='Success',
                    message=(
                        f"Selected docs: {', '.join([dt.label for dt in selected_doc_types])} | "
                        f"Missing: {', '.join(missing_docs_labels)} | Expired: {', '.join(expired_docs_labels)} | "
                        f"CC: {', '.join(cc_emails) if cc_emails else 'None'}"
                    )
                )

            except Exception as e:
                logger.error(f"Error sending email to {tenant.email}: {e}", exc_info=True)
                EmailReminderLog.objects.create(
                    tenant=tenant,
                    email=tenant.email,
                    subject=subject,
                    status='Failed',
                    message=str(e)
                )

        messages.success(request, "Reminders sent to applicable tenants.")
        return redirect('email_reminder_logs')

    messages.error(request, "Invalid request.")
    return redirect('email_reminder_logs')




@permission_required('files.view_emailreminderlog', raise_exception=True)
@login_required(login_url='login')
def email_reminder_logs_view(request):
    query = request.GET.get('q', '')
    status = request.GET.get('status', '')

    logs = EmailReminderLog.objects.select_related('tenant').order_by('-sent_at')

    if query:
        logs = logs.filter(
            Q(tenant__name__icontains=query) | Q(email__icontains=query)
        )
    if status:
        logs = logs.filter(status=status)

    paginator = Paginator(logs, 30)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    all_doc_types = DocumentType.objects.all()  # Add this line to fetch all document types

    return render(request, 'files/email_reminder_logs.html', {
        'page_obj': page_obj,
        'query': query,
        'status_filter': status,
        'all_doc_types': all_doc_types,  # Pass it here to the template
    })

@permission_required('files.delete_emailreminderlog', raise_exception=True)
@login_required(login_url='login')
@require_POST
def delete_all_email_logs(request):
    EmailReminderLog.objects.all().delete()
    messages.success(request, "All email reminder logs have been deleted.")
    return redirect('email_reminder_logs')

@login_required(login_url='login')
@permission_required('files.view_emailreminderlog', raise_exception=True)
def export_email_logs_excel(request):
    # Create workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Email Reminder Logs"

    # Header row
    headers = ['Tenant', 'Email', 'Status', 'Subject', 'Sent At']
    ws.append(headers)

    # Fetch all logs ordered by sent_at descending
    logs = EmailReminderLog.objects.select_related('tenant').order_by('-sent_at')

    for log in logs:
        ws.append([
            log.tenant.name,
            log.email,
            log.status,
            log.subject,
            log.sent_at.strftime('%Y-%m-%d %H:%M')
        ])

    # Prepare HTTP response with Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=email_reminder_logs.xlsx'

    wb.save(response)
    return response



@login_required(login_url='login')
@permission_required('files.add_tenant', raise_exception=True)
def bulk_upload_tenants(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        
        try:
            # Read the Excel file
            if excel_file.name.endswith('.csv'):
                df = pd.read_csv(excel_file)
            else:
                df = pd.read_excel(excel_file)
            
            # Process each row
            for index, row in df.iterrows():
                tenant_name = row.get('tenant_name', '').strip()
                email = row.get('email', '').strip() or None
                tenant_type_code = row.get('tenant_type', '').strip().lower()
                
                if not tenant_name:
                    continue
                
                # Get or create tenant type
                tenant_type, _ = TenantType.objects.get_or_create(
                    code=tenant_type_code,
                    defaults={'label': tenant_type_code.capitalize()}
                )
                
                # Create tenant
                Tenant.objects.get_or_create(
                    name=tenant_name,
                    defaults={
                        'email': email,
                        'tenant_type_fk': tenant_type
                    }
                )
            
            messages.success(request, f'Successfully processed {len(df)} tenants!')
        
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
        
        return redirect('add_tenant')  # Replace with your add tenant URL name
    
    return redirect('add_tenant')

@login_required(login_url='login')
@permission_required('files.view_tenant', raise_exception=True)
def download_tenant_template(request):
    # Create a sample Excel file
    data = {
        'tenant_name': ['Example Tenant 1', 'Example Tenant 2'],
        'email': ['tenant1@example.com', 'tenant2@example.com'],
        'tenant_type': ['kiosk', 'retail']
    }
    df = pd.DataFrame(data)
    
    # Create a BytesIO buffer
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Tenants', index=False)
    
    # Prepare response
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=tenant_upload_template.xlsx'
    return response

from decimal import Decimal
from django.shortcuts import redirect, get_object_or_404
from django.contrib import messages
from .models import Tenant, Unit


@login_required(login_url='login')
@permission_required('files.add_unit', raise_exception=True)
def add_unit(request, tenant_id):
    tenant = get_object_or_404(Tenant, id=tenant_id)
    
    if request.method == "POST":
        unit_id = request.POST.get("unit_id")
        size_value = request.POST.get("size")
        size_type = request.POST.get("unit_type")  # 'sqm' or 'sqft'

        if not size_value:
            messages.error(request, "Size is required.")
            return redirect("tenant_detail", tenant_id=tenant.id)

        # Convert size to Decimal
        size_value = Decimal(size_value)

        # Set the correct field based on unit_type
        if size_type == "sqm":
            size_sqm = size_value
            size_sqft = None
        else:
            size_sqm = None
            size_sqft = size_value

        # Create the unit
        new_unit = Unit.objects.create(
            tenant=tenant,
            unit_id=unit_id,
            size_sqm=size_sqm,
            size_sqft=size_sqft
        )

        messages.success(request, f"Unit {new_unit.unit_id} added successfully.")
        return redirect("tenant_detail", tenant_id=tenant.id)



from decimal import Decimal
from django.shortcuts import redirect, get_object_or_404
from django.contrib import messages
from .models import Unit

@login_required(login_url='login')
@permission_required('files.change_unit', raise_exception=True)
def update_unit(request, unit_id):
    unit = get_object_or_404(Unit, id=unit_id)
    
    if request.method == "POST":
        unit_id_val = request.POST.get("unit_id")
        size_value = request.POST.get("size")
        size_type = request.POST.get("unit_type")

        if not size_value:
            messages.error(request, "Size is required.")
            return redirect("tenant_detail", tenant_id=unit.tenant.id)

        size_value = Decimal(size_value)

        if size_type == "sqm":
            unit.size_sqm = size_value
            unit.size_sqft = None  # will be auto-calculated in save()
        else:
            unit.size_sqm = None  # will be auto-calculated in save()
            unit.size_sqft = size_value

        unit.unit_id = unit_id_val
        unit.save()

        messages.success(request, f"Unit {unit.unit_id} updated successfully.")
        return redirect("tenant_detail", tenant_id=unit.tenant.id)




from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Unit  # Assuming you have a Unit model

@login_required(login_url='login')
@permission_required('files.delete_unit', raise_exception=True)
def delete_unit(request, unit_id):
    unit = get_object_or_404(Unit, id=unit_id)
    tenant_id = unit.tenant.id  # to redirect back to tenant details
    unit.delete()
    messages.success(request, f'Unit {unit.unit_id} has been deleted.')
    return redirect('tenant_detail', tenant_id=tenant_id)

import openpyxl
from django.contrib import messages
from django.shortcuts import redirect
from .models import Tenant, TenantType, Unit

@login_required(login_url='login')
@permission_required('files.add_tenant', raise_exception=True)
def bulk_upload_tenants(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        
        try:
            if excel_file.name.endswith(('.xlsx', '.xls')):
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active
                
                headers = [cell.value for cell in sheet[1]]
                
                for row in sheet.iter_rows(min_row=2):
                    row_data = {headers[i]: cell.value for i, cell in enumerate(row)}
                    
                    tenant_name = str(row_data.get('tenant_name', '')).strip()
                    email = str(row_data.get('email', '')).strip() or None
                    tenant_type_code = str(row_data.get('tenant_type', '')).strip().lower()
                    
                    if not tenant_name:
                        continue
                    
                    tenant_type, _ = TenantType.objects.get_or_create(
                        code=tenant_type_code,
                        defaults={'label': tenant_type_code.capitalize()}
                    )
                    
                    tenant, created = Tenant.objects.get_or_create(
                        name=tenant_name,
                        defaults={
                            'email': email,
                            'tenant_type_fk': tenant_type
                        }
                    )
                    
                    # --- Handle optional units with type ---
                    unit_index = 1
                    while f'unit_{unit_index}' in row_data:
                        unit_name = row_data.get(f'unit_{unit_index}')
                        unit_size = row_data.get(f'size_{unit_index}')
                        unit_type = str(row_data.get(f'unit_{unit_index}_type', 'sqm')).strip().lower()
                        
                        if unit_name and unit_size:
                            try:
                                unit_size = float(unit_size)
                                if unit_type == 'sqft':
                                    # Convert sqft to sqm
                                    unit_size = round(unit_size / 10.7639, 2)
                                # Create unit
                                Unit.objects.get_or_create(
                                    tenant=tenant,
                                    unit_id=unit_name,
                                    defaults={'size_sqm': unit_size}
                                )
                            except ValueError:
                                pass  # skip invalid sizes
                        unit_index += 1
                
                messages.success(request, f'Successfully processed {sheet.max_row - 1} tenants!')
            else:
                messages.error(request, 'Unsupported file format. Please upload .xlsx or .xls file.')
        
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
        
        return redirect('add_tenant')
    
    return redirect('add_tenant')


import openpyxl
from django.http import HttpResponse
from .models import TenantType


@login_required(login_url='login')
@permission_required('files.view_tenant', raise_exception=True)
def download_tenant_template(request):
    # Create a workbook and select active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tenants Template"

    # Basic columns
    headers = ["tenant_name", "email", "tenant_type"]

    # Add optional unit columns (3 example units)
    for i in range(1, 4):
        headers.extend([f"unit_{i}", f"size_{i}", f"unit_{i}_type"])  # unit_id, size, type

    ws.append(headers)

    # Optionally, fill tenant_type options (from DB)
    tenant_types = TenantType.objects.all()
    if tenant_types.exists():
        ws.append(["Example Tenant", "example@email.com", tenant_types.first().code] + [""] * 9)
    else:
        ws.append(["Example Tenant", "example@email.com", "office"] + [""] * 9)

    # Prepare response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="tenant_template.xlsx"'

    wb.save(response)
    return response




@login_required(login_url='login')
@permission_required('files.view_document', raise_exception=True)
@permission_required('files.add_emailreminderlog', raise_exception=True)
def share_documents(request, tenant_id):
    tenant = get_object_or_404(Tenant, id=tenant_id)

    if request.method == "POST":
        subject = request.POST.get("subject", "Shared Documents")
        base_message = request.POST.get("message_share", "")
        receiver_email = request.POST.get("receiver_email")
        cc_emails_raw = request.POST.get("cc_emails_share", "").strip()
        selected_doc_ids = request.POST.getlist('share_docs')

        if not receiver_email:
            messages.error(request, "Please provide a receiver email.")
            return redirect('tenant_detail', tenant_id=tenant.id)

        cc_emails = [email.strip() for email in cc_emails_raw.split(",") if email.strip()]
        documents = Document.objects.filter(id__in=selected_doc_ids)

        if not documents.exists():
            messages.error(request, "No documents selected to share.")
            return redirect('tenant_detail', tenant_id=tenant.id)

        # Compose email content
        doc_list_text = "<br>".join([f"- {doc.doc_type_fk.label}" for doc in documents])
        final_message = (
            f"Please find attached the requested documents, {tenant.name}:<br><br>"
            f"{doc_list_text}<br><br>"
            f"{base_message}"  # optional additional message
        )

        try:
            email = EmailMessage(
                subject,
                final_message,
                None,
                [receiver_email],
                cc=cc_emails if cc_emails else None,
            )
            email.content_subtype = "html"

            # Attach files
            for doc in documents:
                if doc.file:
                    file_path = os.path.join(settings.MEDIA_ROOT, doc.file.name)
                    email.attach_file(file_path)

            email.send(fail_silently=False)

            EmailReminderLog.objects.create(
                tenant=tenant,
                email=receiver_email,
                subject=subject,
                status='Success',
                message=(
                    f"Shared docs: {', '.join([doc.doc_type_fk.label for doc in documents])} | "
                    f"CC: {', '.join(cc_emails) if cc_emails else 'None'}"
                )
            )
            messages.success(request, f"Documents successfully shared with {receiver_email}.")

        except Exception as e:
            logger.error(f"Error sending email to {receiver_email}: {e}", exc_info=True)
            EmailReminderLog.objects.create(
                tenant=tenant,
                email=receiver_email,
                subject=subject,
                status='Failed',
                message=str(e)
            )
            messages.error(request, f"Error sending email: {str(e)}")

        return redirect('tenant_detail', tenant_id=tenant.id)

    messages.error(request, "Invalid request.")
    return redirect('tenant_detail', tenant_id=tenant.id)



from django.shortcuts import get_object_or_404
from django.http import FileResponse, Http404

@login_required(login_url='login')
@permission_required('files.view_document', raise_exception=True)
def download_document(request, document_id):
    # Get the document or return 404 if not found
    doc = get_object_or_404(Document, pk=document_id)

    if not doc.file:
        raise Http404("File not found.")

    try:
        file_path = doc.file.path  # local file path

        if not os.path.exists(file_path):
            raise Http404("File does not exist.")

        # Stream file to browser
        file_handle = open(file_path, 'rb')
        return FileResponse(
            file_handle,
            as_attachment=True,
            filename=os.path.basename(file_path)
        )

    except NotImplementedError:
        # This happens for remote storage (e.g., S3)
        return redirect(doc.file.url)


from django.http import JsonResponse
from django.core.paginator import Paginator
from django.utils import timezone
from datetime import timedelta
from .models import Document


@login_required(login_url='login')
@permission_required('files.view_document', raise_exception=True)
def expiry_data(request):
    today = timezone.now().date()

    # --- Days filter ---
    days = request.GET.get("days", 7)
    try:
        days = int(days)
    except ValueError:
        days = 7
    filter_end_date = today + timedelta(days=days)

    docs = (
        Document.objects.filter(expiry_date__isnull=False)
        .filter(expiry_date__lte=filter_end_date)  # include already expired too
        .select_related("tenant", "tenant__tenant_type_fk", "doc_type_fk")
    )

    # --- Search filter ---
    search = request.GET.get("search", "")
    if search:
        docs = docs.filter(tenant__name__icontains=search)

    # --- Sorting ---
    sort = request.GET.get("sort", "expiry_date")
    sort_map = {
        "tenant": "tenant__name",
        "tenant_type": "tenant__tenant_type_fk__label",
        "doc_type": "doc_type_fk__label",
        "expiry_date": "expiry_date",
    }
    if sort in sort_map:
        docs = docs.order_by(sort_map[sort])

    # --- Pagination ---
    paginator = Paginator(docs, 15)
    page = request.GET.get("page", 1)
    page_obj = paginator.get_page(page)

    # --- Build JSON data ---
    data = []
    for doc in page_obj:
        days_remaining = (doc.expiry_date - today).days if doc.expiry_date else None
        status = "Expired" if days_remaining is not None and days_remaining < 0 else "Active"

        data.append({
            "tenant": doc.tenant.name,
            "tenant_type": doc.tenant.tenant_type_fk.label if doc.tenant.tenant_type_fk else "",
            "doc_type": doc.doc_type_fk.label if doc.doc_type_fk else "",
            "expiry_date": doc.expiry_date.isoformat() if doc.expiry_date else "",
            "days_remaining": days_remaining,
            "status": status,
            "file_view_url": doc.file.url if doc.file else "",
            "file_download_url": f"/download/{doc.id}/" if doc.file else "",
        })

    return JsonResponse({
        "documents": data,
        "has_next": page_obj.has_next(),
        "has_prev": page_obj.has_previous(),
        "page": page_obj.number,
        "total_pages": paginator.num_pages,
    })



from django.shortcuts import render
from django.utils import timezone
from datetime import timedelta
from django.core.paginator import Paginator
from .models import Document


@login_required(login_url='login')
@permission_required('files.view_document', raise_exception=True)
def track_expiry(request):
    today = timezone.now().date()

    # --- Default filter: 7 days ---
    days = request.GET.get("days", 7)
    try:
        days = int(days)
    except ValueError:
        days = 7

    filter_end_date = today + timedelta(days=days)

    documents = Document.objects.filter(
        expiry_date__range=[today, filter_end_date]
    ).select_related("tenant", "tenant__tenant_type_fk", "doc_type_fk")

    # --- Searching by tenant ---
    search_query = request.GET.get("search", "")
    if search_query:
        documents = documents.filter(tenant__name__icontains=search_query)

    # --- Sorting ---
    sort_by = request.GET.get("sort", "expiry_date")  # default sort
    if sort_by in ["tenant", "tenant_type", "doc_type", "expiry_date"]:
        if sort_by == "tenant":
            documents = documents.order_by("tenant__name")
        elif sort_by == "tenant_type":
            documents = documents.order_by("tenant__tenant_type_fk__label")
        elif sort_by == "doc_type":
            documents = documents.order_by("doc_type_fk__label")
        else:
            documents = documents.order_by("expiry_date")

    # --- Pagination (15 per page) ---
    paginator = Paginator(documents, 15)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "documents": page_obj,
        "search_query": search_query,
        "days": days,
        "sort_by": sort_by,
    }
    return render(request, "files/track_expiry.html", context)


from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from io import BytesIO
import openpyxl
import os

from .models import Document

@login_required(login_url='login')
@permission_required('files.view_document', raise_exception=True)
def export_tracked_expiry_excel(request):
    days = int(request.GET.get("days", 7))
    search = request.GET.get("search", "")

    today = timezone.now().date()
    target_date = today + timezone.timedelta(days=days)

    docs = Document.objects.filter(expiry_date__lte=target_date).select_related(
        'tenant', 'tenant__tenant_type_fk', 'doc_type_fk'
    )

    if search:
        docs = docs.filter(tenant__name__icontains=search)

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tracked Expiry"

    # Header row (added Days Remaining ✅)
    ws.append([
        "Tenant Name",
        "Tenant Type",
        "Document Type",
        "Expiry Date",
        "Days Remaining",   # ✅ new column
        "Status",
        "Email",
        "File Name"
    ])

    # Data rows
    for doc in docs:
        status = "Expired" if doc.expiry_date < today else "Active"
        days_remaining = (doc.expiry_date - today).days if doc.expiry_date else None

        ws.append([
            doc.tenant.name,
            doc.tenant.tenant_type_fk.label if doc.tenant.tenant_type_fk else 'No Type',
            doc.doc_type_fk.label if doc.doc_type_fk else 'N/A',
            doc.expiry_date.strftime('%Y-%m-%d') if doc.expiry_date else 'N/A',
            days_remaining if days_remaining is not None else 'N/A',  # ✅ new value
            status,
            doc.tenant.email or '',
            os.path.basename(doc.file.name) if doc.file else ''
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=tracked_expiry.xlsx'
    return response


from .models import ExpiryRule 
from django.http import JsonResponse
from datetime import datetime, timedelta
from .models import Tenant, DocumentType, TenantType # If you have a table like this
from django.http import JsonResponse
from datetime import datetime, timedelta
from .models import Tenant, DocumentType, Document, ExpiryRule


@login_required(login_url='login')
@permission_required('files.view_document', raise_exception=True)
def calculate_expiry(request):
    tenant_id = request.GET.get('tenant')
    doc_type_id = request.GET.get('doc_type')
    commencement_str = request.GET.get('commencement')
    doc_id = request.GET.get('doc_id')

    if not tenant_id or not doc_type_id or not commencement_str or not doc_id:
        return JsonResponse({'expiry_date': ''})

    try:
        tenant = Tenant.objects.get(id=tenant_id)
        doc_type = DocumentType.objects.get(id=doc_type_id)
        doc = Document.objects.get(id=doc_id)
        commencement_date = datetime.strptime(commencement_str, "%Y-%m-%d").date()

        # Save the new commencement date
        doc.commencement_date = commencement_date

        # Lookup expiry rule
        try:
            rule = ExpiryRule.objects.get(
                tenant_type=tenant.tenant_type_fk,
                doc_type=doc_type
            )
            days_to_expiry = rule.days_valid  # <-- corrected field
        except ExpiryRule.DoesNotExist:
            days_to_expiry = 30  # fallback if no rule exists

        # Calculate and save expiry
        expiry_date = commencement_date + timedelta(days=days_to_expiry)
        doc.expiry_date = expiry_date
        doc.save()

        return JsonResponse({'expiry_date': expiry_date.strftime("%Y-%m-%d")})

    except Exception as e:
        return JsonResponse({'expiry_date': '', 'error': str(e)})

    

from django.http import JsonResponse
from datetime import datetime, timedelta
from .models import Document, ExpiryRule

@login_required(login_url='login')
@permission_required('files.change_document', raise_exception=True)
def update_commencement_date(request):
    if request.method != "POST":
        return JsonResponse({'expiry_date': ''})

    doc_id = request.POST.get('doc_id')
    commencement_str = request.POST.get('commencement')

    if not doc_id or not commencement_str:
        return JsonResponse({'expiry_date': ''})

    try:
        # Fetch the document
        doc = Document.objects.get(id=doc_id)

        # Parse and update commencement date
        commencement_date = datetime.strptime(commencement_str, "%Y-%m-%d").date()
        doc.commencement_date = commencement_date

        # Get the expiry rule for this tenant type and document type
        try:
            rule = ExpiryRule.objects.get(
                tenant_type=doc.tenant.tenant_type_fk,
                doc_type=doc.doc_type_fk
            )
            days_to_expiry = rule.days_valid  # <-- corrected field
        except ExpiryRule.DoesNotExist:
            days_to_expiry = 30  # default if no rule found

        # Update expiry date immediately
        doc.expiry_date = commencement_date + timedelta(days=days_to_expiry)
        doc.save()

        return JsonResponse({'expiry_date': doc.expiry_date.strftime('%Y-%m-%d')})

    except Document.DoesNotExist:
        return JsonResponse({'expiry_date': ''})
    except Exception as e:
        return JsonResponse({'expiry_date': '', 'error': str(e)})



from django.contrib.auth.decorators import login_required, permission_required
from django.http import HttpResponse
import openpyxl
from .models import Tenant

@login_required(login_url='login')
@permission_required('files.view_tenant', raise_exception=True)
def export_tenants_excel(request):
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tenants Report"

    # Write headers
    headers = [
        "Tenant Name", "Tenant Type",
        "Units", "Total Size (Units)",
        "Stores", "Total Size (Stores)",
        "Rent Commencement Date", "Escalation Rate (%)"
    ]
    ws.append(headers)

    # Write tenant data
    for tenant in Tenant.objects.prefetch_related("units").all():
        # Separate units into "stores" and "normal units"
        normal_units = [u for u in tenant.units.all() if "store" not in u.unit_id.lower()]
        store_units = [u for u in tenant.units.all() if "store" in u.unit_id.lower()]

        # Prepare values
        tenant_name = tenant.name
        tenant_type = tenant.tenant_type_fk.label if tenant.tenant_type_fk else tenant.tenant_type or "N/A"

        # Units
        units_str = ", ".join(u.unit_id for u in normal_units) if normal_units else ""
        total_units_size = sum(u.size_sqm or 0 for u in normal_units)

        # Stores
        stores_str = ", ".join(u.unit_id for u in store_units) if store_units else ""
        total_stores_size = sum(u.size_sqm or 0 for u in store_units)

        # Rent commencement date and escalation rate
        rent_comm_date = tenant.commencement_date.strftime("%Y-%m-%d") if tenant.commencement_date else ""
        escalation_rate = float(tenant.escalation_rate) if tenant.escalation_rate else 0.0

        # Append row
        ws.append([
            tenant_name, tenant_type,
            units_str, float(total_units_size),
            stores_str, float(total_stores_size),
            rent_comm_date, escalation_rate
        ])

    # Create HTTP response with Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=tenants_report.xlsx'

    wb.save(response)
    return response



from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Document, ArchivedDocument

def archive_documents(request, doc_id):
    document = get_object_or_404(Document, id=doc_id)

    # Archive the document
    ArchivedDocument.objects.create(
        tenant_name=document.tenant.name,
        tenant_type=document.tenant.tenant_type_fk.label if document.tenant.tenant_type_fk else None,
        doc_type=document.doc_type_fk.label if document.doc_type_fk else "Unknown",
        file=document.file,
        upload_date=document.upload_date,
        commencement_date=document.commencement_date,
        expiry_date=document.expiry_date,
    )

    document.delete()
    messages.success(request, "Document archived successfully.")
    return redirect("dashboard")  # Or your document list page



def archived_documents_list(request):
    search_query = request.GET.get("q", "")
    tenant_type_filter = request.GET.get("tenant_type", "")

    # Base queryset
    archived_docs = ArchivedDocument.objects.all()

    # Apply search
    if search_query:
        archived_docs = archived_docs.filter(
            Q(tenant_name__icontains=search_query) |
            Q(doc_type__icontains=search_query)
        )

    # Apply tenant type filter
    if tenant_type_filter:
        archived_docs = archived_docs.filter(tenant_type=tenant_type_filter)

    # Pagination (10 items per page)
    paginator = Paginator(archived_docs.order_by("-archived_at"), 10)
    page = request.GET.get("page")
    docs_page = paginator.get_page(page)

    # Distinct tenant types for filter dropdown
    tenant_types = ArchivedDocument.objects.exclude(tenant_type__isnull=True).values_list("tenant_type", flat=True).distinct()

    return render(request, "files/archived_documents_list.html", {
        "docs_page": docs_page,
        "search_query": search_query,
        "tenant_types": tenant_types,
        "tenant_type_filter": tenant_type_filter,
    })
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import ArchivedDocument, Document, Tenant, DocumentType
from django.utils import timezone

def restore_archived_document(request, archived_id):
    archived = get_object_or_404(ArchivedDocument, id=archived_id)
    
    # Find the tenant object
    tenant = Tenant.objects.filter(name=archived.tenant_name).first()
    doc_type = DocumentType.objects.filter(label=archived.doc_type).first()
    
    if not tenant or not doc_type:
        messages.error(request, "Cannot restore: Tenant or Document Type not found.")
        return redirect('archives')  # replace with your archive page name

    # Create a new Document
    Document.objects.create(
        tenant=tenant,
        doc_type_fk=doc_type,
        file=archived.file,
        upload_date=archived.upload_date,
        commencement_date=archived.commencement_date,
        expiry_date=archived.expiry_date
    )

    # Optionally delete from archive
    archived.delete()

    messages.success(request, "Document restored successfully!")
    return redirect('archived_documents_list')
from django.shortcuts import get_object_or_404, redirect
from django.http import FileResponse, Http404
from django.contrib.auth.decorators import login_required, permission_required
import os
from .models import Document

from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import FileResponse, Http404
import os
from .models import ArchivedDocument  # <-- make sure this is your archived model

@login_required(login_url='login')
@permission_required('files.view_archived_document', raise_exception=True)
def download_archived_document(request, archived_id):
    # Get the archived document from the ArchivedDocument table
    doc = get_object_or_404(ArchivedDocument, pk=archived_id)

    if not doc.file:
        raise Http404("File not found.")

    try:
        file_path = doc.file.path  # local file path

        if not os.path.exists(file_path):
            raise Http404("File does not exist.")

        # Stream file to browser
        file_handle = open(file_path, 'rb')
        return FileResponse(
            file_handle,
            as_attachment=True,
            filename=os.path.basename(file_path)
        )

    except NotImplementedError:
        # For remote storage (like S3)
        return redirect(doc.file.url)

from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from .models import ArchivedDocument  # your archived model

@login_required(login_url='login')
@permission_required('files.delete_archived_document', raise_exception=True)
def delete_archived_document(request, archived_id):
    """
    Deletes an archived document.
    """
    # Get the archived document
    doc = get_object_or_404(ArchivedDocument, pk=archived_id)

    # Delete the file from storage if it exists
    if doc.file and os.path.exists(doc.file.path):
        os.remove(doc.file.path)

    # Delete the record from the database
    doc.delete()

    # Optional: add a success message
    messages.success(request, "Archived document deleted successfully.")

    # Redirect back to the archived documents list
    return redirect('archived_documents_list')  # change to your URL name
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Tenant

# ✅ Update Commencement Date
def update_tenant_date(request, tenant_id):
    tenant = get_object_or_404(Tenant, id=tenant_id)

    if request.method == "POST":
        date_value = request.POST.get("commencement_date")
        if date_value:
            tenant.commencement_date = date_value
            tenant.save()
            messages.success(request, f"Commencement date updated for {tenant.name}")
        else:
            tenant.commencement_date = None
            tenant.save()
            messages.info(request, f"Commencement date cleared for {tenant.name}")

    return redirect("tenant_detail", tenant_id=tenant.id)


# ✅ Update Escalation Rate

@login_required
@permission_required('files.change_tenant', raise_exception=True)
def update_tenant_rate(request, tenant_id):
    tenant = get_object_or_404(Tenant, id=tenant_id)

    if request.method == "POST":
        rate_value = request.POST.get("escalation_rate")
        if rate_value:
            tenant.escalation_rate = rate_value
            tenant.save()
            messages.success(request, f"Escalation rate updated for {tenant.name}")
        else:
            tenant.escalation_rate = None
            tenant.save()
            messages.info(request, f"Escalation rate cleared for {tenant.name}")

    return redirect("tenant_detail", tenant_id=tenant.id)

# views.py
from django.shortcuts import render
from .models import Tenant, Document

from django.shortcuts import render

def data_point_view(request):
    """
    Render the Data Point dashboard with buttons for exporting data.
    No extra context is needed since the page only contains export buttons.
    """
    return render(request, 'files/data_point.html')

from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Tenant
from django.contrib.auth.decorators import login_required, permission_required

from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Tenant
from django.contrib.auth.decorators import login_required, permission_required

@login_required
@permission_required('files.change_tenant', raise_exception=True)
def update_tenant_date_rate(request, tenant_id):
    tenant = get_object_or_404(Tenant, id=tenant_id)

    if request.method == "POST":
        date_value = request.POST.get("commencement_date")
        rate_value = request.POST.get("escalation_rate")

        # If empty, reset the fields
        tenant.commencement_date = date_value if date_value else None
        tenant.escalation_rate = rate_value if rate_value else None

        tenant.save()
        messages.success(request, f"Tenant '{tenant.name}' updated successfully!")

    return redirect("tenant_detail", tenant_id=tenant.id)

